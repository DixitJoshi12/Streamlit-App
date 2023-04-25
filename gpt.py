
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import streamlit.components.v1 as components
from st_aggrid import GridOptionsBuilder, AgGrid, GridUpdateMode, DataReturnMode, JsCode
import numpy as np
import plotly.graph_objects as go

_funct = st.sidebar.radio(label="Master Data", options=[
                          'Mapping', 'Hubble', 'Bridges'])


@st.cache_data
def data_upload(file):
    book = load_workbook(file,
                         data_only=True)
    return book


def show_grid(book, file, headerTitle):

    dfs = {}
    for sheet in book.sheetnames:
        df = pd.read_excel(file, sheet_name=sheet, header=0)

        dfs[sheet] = df

        gb = GridOptionsBuilder.from_dataframe(dfs[sheet])

        gb.configure_pagination(enabled=True,
                                paginationAutoPageSize=False)

        gb.configure_side_bar()
        st.header(headerTitle)
        gridOptions = gb.build()
        AgGrid(df, gridOptions=gridOptions)


def data_visual(df):
    st.header("Data Visualization")
    st.bar_chart(data=df)


if _funct == "Mapping":
    book = data_upload("mapping.xlsx")
    show_grid(book, "mapping.xlsx", "Mapping")

if _funct == "Hubble":
    book = data_upload("hubble.xlsx")
    show_grid(book, "hubble.xlsx", "Hubble Data")

if _funct == "Bridges":
    book = data_upload("bridge.xlsx")
    show_grid(book, "bridge.xlsx", "Bridges Data")
    # st.header("Data Visualization")
    df = pd.read_excel('bridge.xlsx')
    print(df.columns.to_numpy())
    fig = go.Figure(go.Waterfall(
        name="20", orientation="v",
        measure=["absolute", "relative", "relative", "relative", "relative",
                 "relative",  "relative",  "relative",  "relative",  "relative", "total"],

        x=df.columns.to_numpy(),
        textposition="outside",
        # text = ["+60", "+80", "", "-40", "-20", "Total"],
        y=df.values.flatten(),
        base=0,
        decreasing={"marker": {"color": "crimson",
                               "line": {"color": "lightsalmon", "width": 2}}},
        increasing={"marker": {"color": "forestgreen",
                               "line": {"color": "lightgreen", "width": 2}}},
        totals={"marker": {"color": "mediumblue"}},
        connector={"line": {"color": "rgb(63, 63, 63)"}},))
    fig.update_layout(
        title="Variance Analysis",
        showlegend=False)

    fig.show()
    print(df.values.flatten())

    # st.plotly_chart#
